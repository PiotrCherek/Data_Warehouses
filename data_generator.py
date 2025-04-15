import random
from faker import Faker
import datetime
import openpyxl
import unidecode
import os
fakePL = Faker('pl_PL')

#board_game_names = open('board_game_names.txt', 'r').read().split('\n')
#board_games = open('board_games.txt', 'r').read().split('\n')
with open('board_games.txt', 'r') as file:
    board_games = file.readlines()
    for i in range(len(board_games)):
        board_games[i] = board_games[i].rstrip().split()
#print(board_games)
city_districts = open('city_districts.txt', 'r').read().split('\n')

def number_of_records_T1(number):
    return (number * 0.7) - (number * 0.7 % 1)

def random_bool():
    return random.choice([True, False])

def generate_name():
    name = fakePL.first_name()
    return unidecode.unidecode(name)

def generate_surname():
    surname = fakePL.last_name()
    return unidecode.unidecode(surname)

def generate_pesel():
    while True:
        pesel = fakePL.pesel()
        if pesel[0] > '6':
            break
    return pesel

def generate_birth_date():
    date = fakePL.date_of_birth(minimum_age=18, maximum_age=80)
    return date.strftime('%Y-%m-%d')

# WORKERS
workers_T1 = []
workers_T2 = []
number_of_workers_T2 = 10
number_of_workers_T1 = number_of_records_T1(number_of_workers_T2)
for i in range(number_of_workers_T2):
    pesel = generate_pesel()
    name = generate_name()
    surname = generate_surname()
    if i < number_of_workers_T1:
        workers_T1.append({
            'pesel': pesel,
            'name': name,
            'surname': surname
        })
    workers_T2.append({
        'pesel': pesel,
        'name': name,
        'surname': surname
    })

# CUSTOMER ACCOUNT INFO
customers_T1 = []
customers_T2 = []
number_of_customers_T2 = 1000
number_of_customers_T1 = number_of_records_T1(number_of_customers_T2)
starter_customer_code = 10000
for i in range(number_of_customers_T2):
    customer_code = starter_customer_code + i
    name = generate_name()
    surname = generate_surname()
    birth_date = generate_birth_date()
    is_subscribed = random_bool()
    if i < number_of_customers_T1:
        customers_T1.append({
            'customer_code': customer_code,
            'name': name,
            'surname': surname,
            'birth_date': birth_date
        })
    customers_T2.append({
        'customer_code': customer_code,
        'name': name,
        'surname': surname,
        'birth_date': birth_date
    })

def generate_price_pool(min_prize_pool, max_prize_pool):
    random_number = random.randint(min_prize_pool, max_prize_pool)
    if random_number % 100 != 0:
        random_number = random_number - (random_number % 100)
    return random_number

def generate_entry_fee(min_entry_fee, max_entry_fee):
    random_number = random.randint(min_entry_fee, max_entry_fee)
    if random_number % 10 != 0:
        random_number = random_number - (random_number % 10)
    return random_number

# TOURNAMENTS
number_of_tournaments_T2 = 100
number_of_tournaments_T1 = number_of_records_T1(number_of_tournaments_T2)
tournaments_T1 = []
tournaments_T2 = []
current_tournament_id = 1
max_prize_pool = 5000
min_prize_pool = 100
min_entry_fee = 10
max_entry_fee = 100
for i in range(number_of_tournaments_T2):
    tournament_id = current_tournament_id
    game_id = random.randint(1, len(board_games))
    date = fakePL.date_between(start_date='-2y', end_date='today').strftime('%Y-%m-%d')
    prize_pool = generate_price_pool(min_prize_pool, max_prize_pool)
    entry_fee = generate_entry_fee(min_entry_fee, max_entry_fee)
    responsible_worker = random.choice(workers_T1)['pesel']
    if i < number_of_tournaments_T1:
        tournaments_T1.append({
            'tournament_id': tournament_id,
            'game_id': game_id,
            'date': date,
            'prize_pool': prize_pool,
            'entry_fee': entry_fee,
            'responsible_worker': responsible_worker,
            'num_of_winners': random.randint(3, 5)
        })
    tournaments_T2.append({
        'tournament_id': tournament_id,
        'game_id': game_id,
        'date': date,
        'prize_pool': prize_pool,
        'entry_fee': entry_fee,
        'responsible_worker': responsible_worker,
        'num_of_winners': random.randint(3, 5)
    })
    current_tournament_id += 1

def generate_tournament_participants(number_of_participants):
    sample = random.sample(customers_T1, number_of_participants)
    return [customer['customer_code'] for customer in sample]

def split_prize_pool(num_of_splits, prize_pool):
    splits_percentages = []
    for i in range(num_of_splits):
        splits_percentages.append(round(random.randint(1, 10) * 0.1,1))
    while sum(splits_percentages) > 1:
        random_split_index = random.randint(0, num_of_splits - 1)
        if splits_percentages[random_split_index] > 0.1:
            splits_percentages[random_split_index] -= 0.1
            splits_percentages[random_split_index] = round(splits_percentages[random_split_index], 1)
    while sum(splits_percentages) < 1:
        random_split_index = random.randint(0, num_of_splits - 1)
        if splits_percentages[random_split_index] < 1:
            splits_percentages[random_split_index] += 0.1
            splits_percentages[random_split_index] = round(splits_percentages[random_split_index], 1)
    for i in range(num_of_splits):
        splits_percentages[i] = round(prize_pool * splits_percentages[i])
    return sorted(splits_percentages, reverse=True)

def generate_winnings_list(prize_pool, number_of_winners):
    return split_prize_pool(number_of_winners, prize_pool)

# TOURNAMENT PARTICIPANTS
tournament_participants_T1 = []
tournament_participants_T2 = []
for i in range(number_of_tournaments_T2):
    number_of_participants = random.randint(5, 100)
    number_of_winners = tournaments_T2[i]['num_of_winners']
    current_tournament_participants = generate_tournament_participants(number_of_participants)
    winnings_list = generate_winnings_list(tournaments_T2[i]['prize_pool'], number_of_winners)
    for j in range(number_of_participants):
        if i < number_of_tournaments_T1:
            tournament_participants_T1.append({
                'customer_code': current_tournament_participants[j],
                'tournament_id': tournaments_T2[i]['tournament_id'],
                'place': j + 1,
                'prize_amount': winnings_list[j] if j < len(winnings_list) else 0
            })
        tournament_participants_T2.append({
            'customer_code': current_tournament_participants[j],
            'tournament_id': tournaments_T2[i]['tournament_id'],
            'place': j + 1,
            'prize_amount': winnings_list[j] if j < len(winnings_list) else 0
        })

# POSTERS
posters_T1 = []
posters_T2 = []
for i in range(number_of_tournaments_T2):
    number_of_districts = random.randint(2, 7)
    districts = random.sample(city_districts, number_of_districts)
    for j in range(number_of_districts):
        number_of_posters = random.randint(3, 10)
        if i < number_of_tournaments_T1:
            posters_T1.append({
                'tournament_id': tournaments_T2[i]['tournament_id'],
                'district': districts[j],
                'number_of_posters': number_of_posters
            })
        posters_T2.append({
            'tournament_id': tournaments_T2[i]['tournament_id'],
            'district': districts[j],
            'number_of_posters': number_of_posters
        })

# OWNED BOARD GAMES
our_board_games_T1 = []
our_board_games_T2 = []
starter_game_id = 1
for game in board_games:
    our_board_games_T1.append({
        'game_id': starter_game_id,
        'name': game[0],
        'category': game[1],
        'quantity': random.randint(1, 5),
        'rent_price': random.randint(1, 3) * 5
    })
    our_board_games_T2.append({
        'game_id': starter_game_id,
        'name': game[0],
        'category': game[1],
        'quantity': random.randint(1, 5),
        'rent_price': random.randint(1, 3) * 5
    })
    starter_game_id += 1
for i in range(3):
    index = 2*(i+1)
    our_board_games_T2.append({
        'game_id': index+1,
        'name': board_games[index][0],
        'category': board_games[index][1],
        'quantity': random.randint(6, 8),
        'rent_price': random.randint(1, 3) * 5
    })

# RENTS
rents_T1 = []
rents_T2 = []
number_of_rents_T2 = 10000
number_of_rents_T1 = number_of_records_T1(number_of_rents_T2)

starter_rent_id = 1
for i in range(number_of_rents_T2):
    random_customer = random.choice(customers_T1)
    random_game = random.choice(our_board_games_T1)
    date = fakePL.date_between(start_date='-2y', end_date='today').strftime('%Y-%m-%d')
    if i < number_of_rents_T1:
        rents_T1.append({
            'rent_id': starter_rent_id + i,
            'customer_code': random_customer['customer_code'],
            'game': random_game['game_id'],
            'date_of_rent': date
        })
    rents_T2.append({
        'rent_id': starter_rent_id + i,
        'customer_code': random_customer['customer_code'],
        'game': random_game['game_id'],
        'date_of_rent': date
    })

# Sample data generation functions (replace with your actual functions)
def generate_price_pool(min_val, max_val):
    return random.randint(min_val, max_val)

def generate_winnings_list(prize_pool):
    winnings = []
    remaining_pool = prize_pool
    for i in range(3):  # Generate prizes for 1st, 2nd, 3rd places
        if remaining_pool <= 0:
            winnings.append(0)
            continue
        prize = random.randint(0, remaining_pool)
        winnings.append(prize)
        remaining_pool -= prize
    winnings.append(remaining_pool) #4th place
    winnings.append(0)# 5th place
    return winnings

def generate_entry_fee(min_val, max_val):
    return random.randint(min_val, max_val)

# Sample data
fakePL = Faker('pl_PL')
owned_board_games = [
    {'name': 'Chess'},
    {'name': 'Scrabble'},
    {'name': 'Catan'},
    {'name': 'Ticket to Ride'},
    {'name': 'Monopoly'}
]
min_prize_pool = 100
max_prize_pool = 1000
min_entry_fee = 10
max_entry_fee = 50
possible_additional_info = [
    'rents disabled during tournament',
    'morning tournament',
    'evening tournament',
    'meet 30 minutes before start',
    'do not bring children',
    'free snacks and drinks',
    'no additional info'
]
number_of_upcoming_tournaments = 25 

if os.path.exists('upcoming_tournaments_T1.xlsx'):
    # Load existing Excel file
    workbook = openpyxl.load_workbook('upcoming_tournaments_T1.xlsx')
    sheet = workbook.active
    max_row = sheet.max_row
    current_tournament_id = sheet.cell(row=max_row, column=1).value + 1 if max_row > 1 else 1 
    existing_tournaments = []
    headers = [sheet.cell(row=1, column=col).value for col in range(1, sheet.max_column + 1)]
    for row in range(2, max_row + 1):
        tournament = {}
        for col_num, header in enumerate(headers, 1):
            tournament[header] = sheet.cell(row=row, column=col_num).value
        existing_tournaments.append(tournament)

    new_tournaments = []

    while (len(existing_tournaments) + len(new_tournaments) < number_of_upcoming_tournaments):
        random_game = random.choice(owned_board_games)
        prize_pool = generate_price_pool(min_prize_pool, max_prize_pool)
        winnings_list = generate_winnings_list(prize_pool)
        date = fakePL.date_between(start_date=datetime.date.today(), end_date=datetime.date.today() + datetime.timedelta(days=90))
        formatted_date = date.strftime('%Y-%m-%d')
        while len(winnings_list) < 5:
            winnings_list.append(0)
        new_tournaments.append({
            'tournament_id': current_tournament_id,
            'date': formatted_date,
            'name_of_game': random_game['name'],
            'entry_price': generate_entry_fee(min_entry_fee, max_entry_fee),
            'first_place_prize': winnings_list[0],
            'second_place_prize': winnings_list[1],
            'third_place_prize': winnings_list[2],
            'fourth_place_prize': winnings_list[3],
            'fifth_place_prize': winnings_list[4],
            'additional_info': random.choice(possible_additional_info)
        })
        current_tournament_id += 1
    all_tournaments = existing_tournaments + new_tournaments

    # Write to Excel file
    if not existing_tournaments:
        headers = list(all_tournaments[0].keys())
        for col_num, header in enumerate(headers, 1):
            sheet.cell(row=1, column=col_num, value=header)

    for row_num, tournament in enumerate(all_tournaments, 2 if existing_tournaments else 1):
        for col_num, header in enumerate(headers, 1):
            sheet.cell(row=row_num, column=col_num, value=tournament[header])

    workbook.save('upcoming_tournaments_T2.xlsx')
    print("Excel file 'upcoming_tournaments_T2.xlsx' created successfully.")
else:
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    current_tournament_id = 1
    upcoming_tournaments = []

    for i in range(number_of_upcoming_tournaments):
        random_game = random.choice(owned_board_games)
        prize_pool = generate_price_pool(min_prize_pool, max_prize_pool)
        winnings_list = generate_winnings_list(prize_pool)
        date = fakePL.date_between(start_date=datetime.date.today(), end_date=datetime.date.today() + datetime.timedelta(days=60))
        formatted_date = date.strftime('%Y-%m-%d')
        while len(winnings_list) < 5:
            winnings_list.append(0)
        upcoming_tournaments.append({
            'tournament_id': current_tournament_id,
            'date': formatted_date,
            'name_of_game': random_game['name'],
            'entry_price': generate_entry_fee(min_entry_fee, max_entry_fee),
            'first_place_prize': winnings_list[0],
            'second_place_prize': winnings_list[1],
            'third_place_prize': winnings_list[2],
            'fourth_place_prize': winnings_list[3],
            'fifth_place_prize': winnings_list[4],
            'additional_info': random.choice(possible_additional_info)
        })
        current_tournament_id += 1
    # Create Excel file
    headers = list(upcoming_tournaments[0].keys())
    for col_num, header in enumerate(headers, 1):
        sheet.cell(row=1, column=col_num, value=header)
    for row_num, tournament in enumerate(upcoming_tournaments, 2):
        for col_num, header in enumerate(headers, 1):
            sheet.cell(row=row_num, column=col_num, value=tournament[header])
    workbook.save('upcoming_tournaments_T1.xlsx')
    print("Excel file 'upcoming_tournaments_T1.xlsx' created successfully.")
def push_to_file(file_name, data):
    with open(file_name, 'w') as file:
        for row in data:
            line = ''
            for key, value in row.items():
                line += str(value) + '|'
            file.write(line[:-1] + '\n')

push_to_file('customers_T1.bulk', customers_T1)
push_to_file('customers_T2.bulk', customers_T2)
push_to_file('tournaments_T1.bulk', tournaments_T1)
push_to_file('tournaments_T2.bulk', tournaments_T2)
push_to_file('tournament_participants_T1.bulk', tournament_participants_T1)
push_to_file('tournament_participants_T2.bulk', tournament_participants_T2)
# push_to_file('posters_T1.bulk', posters_T1)
# push_to_file('posters_T2.bulk', posters_T2)
push_to_file('owned_board_games_T1.bulk', our_board_games_T1)
push_to_file('owned_board_games_T2.bulk', our_board_games_T2)
push_to_file('rents_T1.bulk', rents_T1)
push_to_file('rents_T2.bulk', rents_T2)
push_to_file('workers_T1.bulk', workers_T1)
push_to_file('workers_T2.bulk', workers_T2)